import io
from datetime import date, timedelta
from fastapi import APIRouter, Depends, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session

from database import get_db
from models import Attendance, Trainee
from dependencies import get_current_admin

router = APIRouter(prefix="/api/v1/reports", tags=["reports"])


def get_attendance_data(db: Session, from_date: date, to_date: date) -> list[dict]:
    records = (
        db.query(Attendance)
        .filter(Attendance.date >= from_date, Attendance.date <= to_date)
        .order_by(Attendance.date.asc(), Attendance.checkin_time.asc())
        .all()
    )

    rows = []
    for r in records:
        trainee = db.query(Trainee).filter(Trainee.id == r.trainee_id).first()
        rows.append({
            "Date": str(r.date),
            "Name": trainee.unique_name if trainee else "Unknown",
            "Check-in": r.checkin_time.strftime("%H:%M:%S") if r.checkin_time else "",
            "Check-out": r.checkout_time.strftime("%H:%M:%S") if r.checkout_time else "",
            "Status": r.status or "",
        })
    return rows


@router.get("/export")
async def export_report(
    format: str = Query("excel", regex="^(excel|pdf)$"),
    from_date: date = Query(None, alias="from"),
    to_date: date = Query(None, alias="to"),
    db: Session = Depends(get_db),
    _admin: dict = Depends(get_current_admin),
):
    if not from_date:
        from_date = date.today() - timedelta(days=date.today().weekday())
    if not to_date:
        to_date = date.today()

    rows = get_attendance_data(db, from_date, to_date)

    if format == "excel":
        return generate_excel(rows, from_date, to_date)
    return generate_pdf(rows, from_date, to_date)


def generate_excel(rows: list[dict], from_date: date, to_date: date) -> StreamingResponse:
    from openpyxl import Workbook
    from openpyxl.styles import Font
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Attendance Report"

    headers = ["Name", "Date", "Check-in", "Check-out", "Status"]
    ws.append(headers)
    bold = Font(bold=True)
    for col_idx in range(1, len(headers) + 1):
        ws.cell(row=1, column=col_idx).font = bold

    for row in rows:
        ws.append([row[h] for h in headers])

    for col_idx in range(1, len(headers) + 1):
        max_len = max(
            len(str(ws.cell(row=r, column=col_idx).value or ""))
            for r in range(1, ws.max_row + 1)
        )
        ws.column_dimensions[get_column_letter(col_idx)].width = max_len + 3

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    fn = f"attendance_{from_date.strftime('%Y%m%d')}_{to_date.strftime('%Y%m%d')}.xlsx"
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={fn}"},
    )


def generate_pdf(rows: list[dict], from_date: date, to_date: date) -> StreamingResponse:
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    styles = getSampleStyleSheet()
    elements = []

    elements.append(Paragraph(f"Attendance Report: {from_date} to {to_date}", styles["Title"]))
    elements.append(Spacer(1, 20))

    headers = ["Name", "Date", "Check-in", "Check-out", "Status"]
    if rows:
        table_data = [headers] + [[row[h] for h in headers] for row in rows]
    else:
        table_data = [headers, ["No records", "", "", "", ""]]

    table = Table(table_data)

    style_cmds = [
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1f2937")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("BOTTOMPADDING", (0, 0), (-1, 0), 12),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f9fafb")]),
    ]

    status_colors = {
        "present": colors.HexColor("#dcfce7"),
        "late": colors.HexColor("#fef9c3"),
        "absent": colors.HexColor("#fee2e2"),
    }
    for i, row in enumerate(rows):
        bg = status_colors.get(row.get("Status", "").lower())
        if bg:
            style_cmds.append(("BACKGROUND", (0, i + 1), (-1, i + 1), bg))

    table.setStyle(TableStyle(style_cmds))
    elements.append(table)

    doc.build(elements)
    buffer.seek(0)

    fn = f"attendance_{from_date.strftime('%Y%m%d')}_{to_date.strftime('%Y%m%d')}.pdf"
    return StreamingResponse(
        buffer,
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename={fn}"},
    )
